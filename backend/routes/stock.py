import uuid
import io
import csv
import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, g, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from extensions import db
from models.stock_model import StockEntry
from utils.auth_middleware import authenticate
from utils.permission_middleware import require_permission
from utils.audit_logger import log_action
from utils.cache_helper import get_cache_manager, invalidate_stock_cache
from utils.helpers import title_case
from utils.rate_limiter import rate_limit

logger = logging.getLogger(__name__)

stock_bp = Blueprint('stock', __name__)

# Cache timeout for stock list (5 minutes - cache is invalidated on stock changes anyway)
# OPTIMIZED: Increased from 2 min to 5 min since we have proper cache invalidation
STOCK_CACHE_TIMEOUT = 300


def _to_num(value, default=None):
    """Parse form input to float. Empty / whitespace / invalid → default (usually None)."""
    if value is None:
        return default
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == '':
            return default
        try:
            return float(stripped)
        except ValueError:
            return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value, default=None):
    """Parse form input to int. Empty / whitespace / invalid → default."""
    if value is None:
        return default
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == '':
            return default
        try:
            return int(float(stripped))  # tolerate "10.0" etc.
        except ValueError:
            return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _to_str_or_none(value):
    """Strip string; empty → None. Non-string values pass through."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def generate_item_code(client_id, product_name):
    """
    Auto-generate item code based on product name and client
    Format: {PRODUCT_INITIALS}-{CLIENT_PREFIX}-{SEQUENCE}
    Example: LAP-550-001 for "Laptop" product
    """
    import re

    # Extract product initials (first 3 letters, alphanumeric only)
    clean_name = re.sub(r'[^a-zA-Z0-9]', '', product_name)
    product_prefix = clean_name[:3].upper() if clean_name else 'ITM'

    # Extract client prefix (first 3 chars of client_id)
    client_prefix = client_id[:3].upper()

    # Find the highest existing sequence for this prefix in one MAX() query,
    # replacing the old while-loop that could issue up to 100 DB queries.
    from sqlalchemy import func

    prefix = f"{product_prefix}-{client_prefix}-"
    existing_max = db.session.query(
        func.max(StockEntry.item_code)
    ).filter(
        StockEntry.client_id == client_id,
        StockEntry.item_code.like(f"{prefix}%"),
    ).scalar()

    if existing_max:
        try:
            current_seq = int(existing_max.rsplit('-', 1)[-1])
            next_sequence = current_seq + 1
        except (ValueError, IndexError):
            # Suffix isn't a plain integer (e.g. UUID fallback) — start from count+1
            next_sequence = (db.session.query(func.count(StockEntry.product_id))
                             .filter_by(client_id=client_id).scalar() or 0) + 1
    else:
        next_sequence = 1

    item_code = f"{prefix}{next_sequence:03d}"

    # Single-query collision guard (covers concurrent inserts at the exact same sequence)
    if StockEntry.query.filter_by(client_id=client_id, item_code=item_code).first():
        item_code = f"{prefix}{uuid.uuid4().hex[:6].upper()}"

    return item_code


@stock_bp.route('', methods=['POST'])
@authenticate
@rate_limit(max_requests=30, window_seconds=60, key_func=lambda: g.user['user_id'])
@require_permission('add_product')
def add_stock():
    """
    Add stock entry with client_id
    Auto-sum if product already exists for this client
    """
    try:
        data = request.get_json() or {}
        client_id = g.user['client_id']

        # ── Required fields ──────────────────────────────────────────────────
        product_name_raw = data.get('product_name')
        if not product_name_raw or not str(product_name_raw).strip():
            return jsonify({'error': 'Missing required field: product_name'}), 400

        quantity = _to_int(data.get('quantity'))
        if quantity is None:
            return jsonify({'error': 'Missing or invalid required field: quantity'}), 400
        if quantity < 0:
            return jsonify({'error': 'quantity must be zero or positive'}), 400

        rate = _to_num(data.get('rate'))
        if rate is None:
            return jsonify({'error': 'Missing or invalid required field: rate'}), 400
        if rate < 0:
            return jsonify({'error': 'rate must be zero or positive'}), 400

        product_name = title_case(str(product_name_raw).strip())
        category = title_case(str(data.get('category') or 'Other').strip() or 'Other')

        # ── Optional numeric fields (empty/blank → None) ─────────────────────
        cost_price      = _to_num(data.get('cost_price'))
        mrp             = _to_num(data.get('mrp'))
        pricing         = _to_num(data.get('pricing'))
        gst_percentage  = _to_num(data.get('gst_percentage'), default=0)
        low_stock_alert = _to_int(data.get('low_stock_alert'), default=10)

        # ── Optional string fields (empty → None) ────────────────────────────
        unit     = _to_str_or_none(data.get('unit')) or 'pcs'
        hsn_code = _to_str_or_none(data.get('hsn_code'))
        barcode  = _to_str_or_none(data.get('barcode'))   # optional — no auto-gen
        item_code_in = _to_str_or_none(data.get('item_code'))

        # Added-by label (v19) — user-typed name for shared-login attribution
        added_by_label = (data.get('added_by_label') or '').strip() or None

        # Apparel / Legal Metrology (v13) — all optional
        apparel = {
            'brand_name':          _to_str_or_none(data.get('brand_name')),
            'size_variant':        _to_str_or_none(data.get('size_variant')),
            'colour':              _to_str_or_none(data.get('colour')),
            'country_of_origin':   _to_str_or_none(data.get('country_of_origin')),
            'manufacture_date':    _to_str_or_none(data.get('manufacture_date')),
            'importer_name':       _to_str_or_none(data.get('importer_name')),
            'importer_address':    _to_str_or_none(data.get('importer_address')),
            'consumer_care_phone': _to_str_or_none(data.get('consumer_care_phone')),
            'consumer_care_email': _to_str_or_none(data.get('consumer_care_email')),
        }

        # ── Find existing product for auto-sum ───────────────────────────────
        existing_product = StockEntry.query.filter_by(
            client_id=client_id,
            product_name=product_name
        ).first()

        if existing_product:
            old_data = existing_product.to_dict()
            existing_product.quantity += quantity
            existing_product.rate = rate
            if cost_price is not None:      existing_product.cost_price = cost_price
            if mrp is not None:             existing_product.mrp = mrp
            if pricing is not None:         existing_product.pricing = pricing
            existing_product.category = category
            existing_product.unit = unit
            existing_product.low_stock_alert = low_stock_alert
            existing_product.gst_percentage = gst_percentage
            if hsn_code is not None:        existing_product.hsn_code = hsn_code

            # item_code: explicit > existing > auto-generated
            if item_code_in:
                existing_product.item_code = item_code_in
            elif not existing_product.item_code:
                existing_product.item_code = generate_item_code(client_id, existing_product.product_name)

            # barcode: set to new value if provided, else preserve existing (don't clear)
            if barcode is not None:
                existing_product.barcode = barcode

            # apparel: overwrite only when a non-null value is supplied
            for field, value in apparel.items():
                if value is not None:
                    setattr(existing_product, field, value)

            existing_product.updated_at = datetime.utcnow()
            db.session.commit()
            invalidate_stock_cache(client_id)
            log_action('UPDATE', 'stock_entry', existing_product.product_id, old_data, existing_product.to_dict())

            return jsonify({
                'success': True,
                'product_id': existing_product.product_id,
                'message': 'Stock updated successfully (quantity added)',
                'product': existing_product.to_dict()
            }), 200

        # ── New product ──────────────────────────────────────────────────────
        item_code = item_code_in or generate_item_code(client_id, product_name)

        new_product = StockEntry(
            product_id=str(uuid.uuid4()),
            client_id=client_id,
            product_name=product_name,
            category=category,
            quantity=quantity,
            rate=rate,
            cost_price=cost_price,
            mrp=mrp,
            pricing=pricing,
            unit=unit,
            low_stock_alert=low_stock_alert,
            item_code=item_code,
            barcode=barcode,
            gst_percentage=gst_percentage,
            hsn_code=hsn_code,
            created_at=datetime.utcnow(),
            created_by=g.user['user_id'],
            added_by_label=added_by_label,
            **apparel,
        )

        db.session.add(new_product)
        db.session.commit()
        invalidate_stock_cache(client_id)
        log_action('CREATE', 'stock_entry', new_product.product_id, None, new_product.to_dict())

        return jsonify({
            'success': True,
            'product_id': new_product.product_id,
            'message': 'Stock added successfully',
            'product': new_product.to_dict()
        }), 201

    except Exception:
        db.session.rollback()
        logger.exception(f"[Stock POST] Failed for client {g.user.get('client_id')}")
        return jsonify({'error': 'Failed to add stock'}), 500


@stock_bp.route('', methods=['GET'])
@authenticate
@require_permission('view_stock')
def get_stock():
    """List stock entries filtered by client_id - OPTIMIZED with caching"""
    try:
        client_id = g.user['client_id']
        category  = request.args.get('category', '')
        search    = request.args.get('search', '')
        # HIGH-2: Cap at MAX_STOCK_LIMIT — limit=0 (falsy) previously skipped the LIMIT clause,
        # loading the entire catalog into memory on every cache-miss.
        MAX_STOCK_LIMIT = 5000
        requested_limit = request.args.get('limit', 0, type=int)
        limit = min(requested_limit, MAX_STOCK_LIMIT) if requested_limit > 0 else MAX_STOCK_LIMIT

        cache     = get_cache_manager()
        cache_key = f"stock:list:{client_id}:{category}:{search}:{limit}"
        cached    = cache.get(cache_key)
        if cached is not None:
            return jsonify(cached), 200

        query = StockEntry.query.filter_by(client_id=client_id)
        if category:
            query = query.filter_by(category=category)
        if search:
            query = query.filter(StockEntry.product_name.ilike(f'%{search}%'))
        query = query.order_by(StockEntry.product_name).limit(limit)

        stock_data = [entry.to_dict() for entry in query.all()]
        response   = {'success': True, 'stock': stock_data}
        cache.set(cache_key, response, STOCK_CACHE_TIMEOUT)
        return jsonify(response), 200

    except Exception as e:
        return jsonify({'error': 'Failed to fetch stock', 'message': str(e)}), 500


@stock_bp.route('/alerts', methods=['GET'])
@authenticate
@require_permission('view_low_stock_alerts')
def get_low_stock_alerts():
    """Get low stock alerts filtered by client_id with full product details"""
    try:
        client_id = g.user['client_id']

        cache = get_cache_manager()
        cache_key = f"stock:alerts:{client_id}"
        cached = cache.get(cache_key)
        if cached is not None:
            return jsonify(cached), 200

        # Get products where quantity <= low_stock_alert
        low_stock = StockEntry.query.filter(
            StockEntry.client_id == client_id,
            StockEntry.quantity <= StockEntry.low_stock_alert
        ).order_by(StockEntry.quantity).all()

        # MED-2: Serialize once via to_dict(); derive compact 'alerts' list from the same data.
        # Previously: to_dict() called once for low_stock_products AND a separate list comprehension
        # for alerts — iterating the ORM objects twice.
        low_stock_dicts = [item.to_dict() for item in low_stock]
        result = {
            'success': True,
            'alerts': [
                {
                    'product_id': d['product_id'],
                    'product_name': d['product_name'],
                    'current_quantity': d['quantity'],
                    'alert_threshold': d['low_stock_alert'],
                    'unit': d['unit']
                }
                for d in low_stock_dicts
            ],
            'low_stock_products': low_stock_dicts,
            'alert_count': len(low_stock_dicts)
        }
        cache.set(cache_key, result, 60)
        return jsonify(result), 200

    except Exception as e:
        return jsonify({'error': 'Failed to fetch alerts', 'message': str(e)}), 500


@stock_bp.route('/<product_id>', methods=['PUT'])
@authenticate
@require_permission('edit_product_details')
def update_stock(product_id):
    """Update stock entry with client_id validation"""
    try:
        client_id = g.user['client_id']
        data = request.get_json()

        # Find product
        product = StockEntry.query.filter_by(
            product_id=product_id,
            client_id=client_id
        ).first()

        if not product:
            return jsonify({'error': 'Product not found'}), 404

        # Store old data for audit
        old_data = product.to_dict()

        # Text fields
        if 'product_name' in data:
            new_name = _to_str_or_none(data['product_name'])
            if new_name:
                product.product_name = title_case(new_name)
        if 'category' in data:
            new_cat = _to_str_or_none(data['category'])
            product.category = title_case(new_cat) if new_cat else product.category
        if 'unit' in data:
            new_unit = _to_str_or_none(data['unit'])
            if new_unit:
                product.unit = new_unit
        if 'hsn_code' in data:
            product.hsn_code = _to_str_or_none(data['hsn_code'])

        # Numeric fields:
        #  - Required (quantity, rate): reject empty/invalid with 400.
        #  - Optional (cost_price, mrp, pricing): empty → leave unchanged
        #    (avoids silent data loss when the form re-POSTs blank optional fields).
        if 'quantity' in data:
            qty = _to_int(data['quantity'])
            if qty is None or qty < 0:
                return jsonify({'error': 'Invalid quantity'}), 400
            product.quantity = qty
        if 'rate' in data:
            rate = _to_num(data['rate'])
            if rate is None or rate < 0:
                return jsonify({'error': 'Invalid rate'}), 400
            product.rate = rate
        if 'cost_price' in data and data['cost_price'] not in (None, ''):
            product.cost_price = _to_num(data['cost_price'])
        if 'mrp' in data and data['mrp'] not in (None, ''):
            product.mrp = _to_num(data['mrp'])
        if 'pricing' in data and data['pricing'] not in (None, ''):
            product.pricing = _to_num(data['pricing'])
        if 'low_stock_alert' in data:
            product.low_stock_alert = _to_int(data['low_stock_alert'], default=product.low_stock_alert)
        if 'gst_percentage' in data:
            product.gst_percentage = _to_num(data['gst_percentage'], default=0)

        # item_code: keep old if empty, auto-generate if still missing
        if 'item_code' in data:
            new_code = _to_str_or_none(data['item_code'])
            if new_code:
                product.item_code = new_code
        if not product.item_code:
            product.item_code = generate_item_code(client_id, product.product_name)

        # barcode: optional — empty string clears it to NULL
        if 'barcode' in data:
            product.barcode = _to_str_or_none(data['barcode'])

        # Apparel / Legal Metrology fields (all optional strings)
        for apparel_field in (
            'brand_name', 'size_variant', 'colour', 'country_of_origin',
            'manufacture_date', 'importer_name', 'importer_address',
            'consumer_care_phone', 'consumer_care_email',
        ):
            if apparel_field in data:
                setattr(product, apparel_field, _to_str_or_none(data[apparel_field]))

        product.updated_at = datetime.utcnow()

        db.session.commit()

        # Invalidate stock cache so changes appear immediately
        invalidate_stock_cache(client_id)

        # Log action
        log_action('UPDATE', 'stock_entry', product_id, old_data, product.to_dict())

        return jsonify({
            'success': True,
            'message': 'Stock updated successfully',
            'product': product.to_dict()
        }), 200

    except Exception:
        db.session.rollback()
        logger.exception(f"[Stock PUT] Failed for product {product_id}")
        return jsonify({'error': 'Failed to update stock'}), 500


@stock_bp.route('/<product_id>', methods=['DELETE'])
@authenticate
@require_permission('delete_product')
def delete_stock(product_id):
    """Delete stock entry with client_id validation.

    Uses a raw DBAPI connection to completely bypass the SQLAlchemy ORM session.
    The ORM's BranchInventory backref causes an automatic SET NULL on
    branch_inventory.product_id during session flush, which violates the
    NOT NULL constraint in PostgreSQL.
    """
    try:
        client_id = g.user['client_id']
        pid = str(product_id)
        cid = str(client_id)

        # Get the raw DBAPI connection — completely bypasses ORM identity map & autoflush
        is_sqlite = str(db.engine.url).startswith('sqlite')
        ph = '?' if is_sqlite else '%s'  # placeholder syntax differs per driver

        conn = db.engine.raw_connection()
        try:
            cur = conn.cursor()

            # Fetch product data for audit log
            cur.execute(
                f"SELECT product_name, category, rate, quantity, unit, item_code "
                f"FROM stock_entry WHERE product_id = {ph} AND client_id = {ph}",
                (pid, cid)
            )
            row = cur.fetchone()

            if not row:
                conn.close()
                return jsonify({'error': 'Product not found'}), 404

            old_data = {
                'product_id': pid,
                'product_name': row[0],
                'category': row[1],
                'rate': float(row[2]) if row[2] else None,
                'quantity': row[3],
                'unit': row[4],
                'item_code': row[5],
            }

            # Delete all child rows, then the parent
            cur.execute(f"DELETE FROM branch_inventory WHERE product_id = {ph}", (pid,))
            cur.execute(f"DELETE FROM stock_transfer_items WHERE product_id = {ph}", (pid,))
            cur.execute(f"DELETE FROM bulk_stock_order_item WHERE product_id = {ph}", (pid,))
            cur.execute(f"UPDATE supplier_delivery_items SET product_id = NULL WHERE product_id = {ph}", (pid,))
            cur.execute(f"DELETE FROM stock_entry WHERE product_id = {ph} AND client_id = {ph}", (pid, cid))

            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

        # Evict any cached ORM objects for this product so they don't linger
        db.session.expire_all()

        # Invalidate stock cache so deletion appears immediately
        invalidate_stock_cache(client_id)

        # Log action
        log_action('DELETE', 'stock_entry', product_id, old_data, None)

        return jsonify({
            'success': True,
            'message': 'Stock deleted successfully'
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete stock', 'message': str(e)}), 500


@stock_bp.route('/bulk-import', methods=['POST'])
@authenticate
@require_permission('import_stock')
@rate_limit(max_requests=5, window_seconds=60, key_func=lambda: g.user['user_id'], error_message='Bulk import is limited to 5 times per minute.')
def bulk_import_stock():
    """
    Bulk import stock from CSV or Excel file
    Supports both create new and update existing products
    """
    try:
        client_id = g.user['client_id']

        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Check file extension
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        if file_ext not in ['csv', 'xlsx', 'xls']:
            return jsonify({'error': 'Invalid file format. Only CSV, XLSX, XLS files are allowed'}), 400

        # Read file based on extension
        try:
            if file_ext == 'csv':
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': 'Failed to read file', 'message': str(e)}), 400

        # Read default added_by_label from form data (dialog-level default)
        default_added_by_label = (request.form.get('default_added_by_label') or '').strip() or None

        # Validate required columns
        required_columns = ['product_name', 'quantity', 'rate']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return jsonify({
                'error': f'Missing required columns: {", ".join(missing_columns)}',
                'required_columns': required_columns,
                'found_columns': list(df.columns)
            }), 400

        # ==================== PRE-LOAD: eliminates N+1 per-row DB lookups ====================
        # One query loads all existing stock for this client into dicts.
        # The inner loop never hits the DB for product-name or barcode lookups.
        _all_existing = StockEntry.query.filter_by(client_id=client_id).all()
        existing_by_name = {e.product_name: e for e in _all_existing}
        existing_barcodes = {e.barcode: e.product_id for e in _all_existing if e.barcode}
        existing_item_codes = {e.item_code for e in _all_existing if e.item_code}

        # Per-prefix sequence counters for item_code generation (avoids repeated MAX queries).
        import re as _re_bulk
        _prefix_seq: dict = {}

        def _gen_item_code_fast(pname: str) -> str:
            """Generate a unique item code without any DB round-trips."""
            clean = _re_bulk.sub(r'[^a-zA-Z0-9]', '', pname)
            prod_pfx = clean[:3].upper() if clean else 'ITM'
            cli_pfx = client_id[:3].upper()
            prefix = f"{prod_pfx}-{cli_pfx}-"
            if prefix not in _prefix_seq:
                # Seed counter from the highest existing code with this prefix.
                candidates = [c for c in existing_item_codes if c and c.startswith(prefix)]
                if candidates:
                    try:
                        max_seq = max(
                            int(c.rsplit('-', 1)[-1]) for c in candidates
                            if c.rsplit('-', 1)[-1].isdigit()
                        )
                        _prefix_seq[prefix] = max_seq + 1
                    except (ValueError, IndexError):
                        _prefix_seq[prefix] = len(candidates) + 1
                else:
                    _prefix_seq[prefix] = 1
            seq = _prefix_seq[prefix]
            _prefix_seq[prefix] += 1
            code = f"{prefix}{seq:03d}"
            if code in existing_item_codes:
                code = f"{prefix}{uuid.uuid4().hex[:6].upper()}"
            existing_item_codes.add(code)
            return code

        # Process each row
        success_count = 0
        error_count = 0
        updated_count = 0
        created_count = 0
        errors = []

        for index, row in df.iterrows():
            savepoint = db.session.begin_nested()
            try:
                # Skip rows with missing required fields
                if pd.isna(row['product_name']) or pd.isna(row['quantity']) or pd.isna(row['rate']):
                    savepoint.rollback()
                    error_count += 1
                    errors.append(f"Row {index + 2}: Missing required fields")
                    continue

                product_name = title_case(str(row['product_name']).strip())
                quantity = int(row['quantity'])
                rate = float(row['rate'])
                category = title_case(str(row['category']).strip()) if 'category' in row and not pd.isna(row['category']) else 'Other'
                unit = str(row['unit']).strip() if 'unit' in row and not pd.isna(row['unit']) else 'pcs'
                low_stock_alert = int(row['low_stock_alert']) if 'low_stock_alert' in row and not pd.isna(row['low_stock_alert']) else 10
                # Support both 'purchase_price' and 'cost_price' column names (purchase_price takes priority)
                cost_price = None
                if 'purchase_price' in row and not pd.isna(row['purchase_price']):
                    cost_price = float(row['purchase_price'])
                elif 'cost_price' in row and not pd.isna(row['cost_price']):
                    cost_price = float(row['cost_price'])
                mrp = float(row['mrp']) if 'mrp' in row and not pd.isna(row['mrp']) else None

                # Handle item_code — use fast in-memory generator (no DB queries)
                item_code = str(row['item_code']).strip() if 'item_code' in row and not pd.isna(row['item_code']) else ''
                if not item_code:
                    item_code = _gen_item_code_fast(product_name)

                # Handle barcode - set to None if empty
                barcode = str(row['barcode']).strip() if 'barcode' in row and not pd.isna(row['barcode']) else ''
                barcode = barcode if barcode else None

                # Handle GST percentage and HSN code
                gst_percentage = float(row['gst_percentage']) if 'gst_percentage' in row and not pd.isna(row['gst_percentage']) else 0
                hsn_code = str(row['hsn_code']).strip() if 'hsn_code' in row and not pd.isna(row['hsn_code']) else ''

                # Resolve added_by_label: row value > dialog default; reject if both missing
                row_label = str(row['added_by_label']).strip() if 'added_by_label' in row and not pd.isna(row.get('added_by_label')) else ''
                row_added_by_label = row_label or default_added_by_label
                if not row_added_by_label:
                    savepoint.rollback()
                    error_count += 1
                    errors.append(f"Row {index + 2}: added_by_label is required (fill the column or provide a dialog default)")
                    continue

                # Validate quantity and rate
                if quantity < 0 or rate < 0:
                    savepoint.rollback()
                    error_count += 1
                    errors.append(f"Row {index + 2}: Quantity and rate must be positive")
                    continue

                # All lookups use pre-loaded dicts — no DB round-trips inside the loop.
                existing_product = existing_by_name.get(product_name)

                # Resolve barcode conflicts using pre-loaded barcodes dict.
                if barcode:
                    owner_id = existing_product.product_id if existing_product else None
                    if existing_barcodes.get(barcode) not in (None, owner_id):
                        errors.append(f"Row {index + 2}: Barcode '{barcode}' already in use — imported '{product_name}' without barcode")
                        barcode = None

                if existing_product:
                    # Update existing product (auto-sum quantity)
                    old_data = existing_product.to_dict()
                    existing_product.quantity += quantity
                    existing_product.rate = rate
                    existing_product.cost_price = cost_price if cost_price is not None else existing_product.cost_price
                    existing_product.mrp = mrp if mrp is not None else existing_product.mrp
                    existing_product.category = category
                    existing_product.unit = unit
                    existing_product.low_stock_alert = low_stock_alert
                    existing_product.gst_percentage = gst_percentage
                    existing_product.hsn_code = hsn_code

                    if not existing_product.item_code:
                        existing_product.item_code = item_code

                    if barcode:
                        existing_product.barcode = barcode

                    existing_product.updated_at = datetime.utcnow()

                    log_action('UPDATE', 'stock_entry', existing_product.product_id, old_data, existing_product.to_dict())
                    # Keep barcode dict consistent for later rows in the same CSV.
                    if barcode:
                        existing_barcodes[barcode] = existing_product.product_id
                    updated_count += 1
                else:
                    new_product = StockEntry(
                        product_id=str(uuid.uuid4()),
                        client_id=client_id,
                        product_name=product_name,
                        category=category,
                        quantity=quantity,
                        rate=rate,
                        cost_price=cost_price,
                        mrp=mrp,
                        unit=unit,
                        low_stock_alert=low_stock_alert,
                        item_code=item_code,
                        barcode=barcode,
                        gst_percentage=gst_percentage,
                        hsn_code=hsn_code,
                        created_at=datetime.utcnow(),
                        created_by=g.user['user_id'],
                        added_by_label=row_added_by_label,
                    )
                    db.session.add(new_product)
                    log_action('CREATE', 'stock_entry', new_product.product_id, None, new_product.to_dict())
                    # Track new product in dicts so duplicate rows later in the CSV are handled.
                    existing_by_name[product_name] = new_product
                    if barcode:
                        existing_barcodes[barcode] = new_product.product_id
                    created_count += 1

                savepoint.commit()
                success_count += 1

            except Exception as e:
                savepoint.rollback()
                error_count += 1
                errors.append(f"Row {index + 2}: {str(e)}")

        # Commit all changes
        db.session.commit()

        # Invalidate stock cache so all imported products appear immediately
        invalidate_stock_cache(client_id)

        return jsonify({
            'success': True,
            'message': 'Bulk import completed',
            'summary': {
                'total_rows': len(df),
                'success_count': success_count,
                'created_count': created_count,
                'updated_count': updated_count,
                'error_count': error_count,
                'errors': errors[:10] if errors else []  # Return first 10 errors
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to import stock', 'message': str(e)}), 500


@stock_bp.route('/download-template', methods=['POST'])
@authenticate
def download_template():
    """Download CSV/Excel template for bulk import with sample data"""
    try:
        import os
        data = request.get_json() or {}
        file_format = data.get('format', 'csv')

        # Get the path to template files
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')

        if file_format == 'csv':
            template_path = os.path.join(template_dir, 'stock_template.csv')
            if os.path.exists(template_path):
                return send_file(
                    template_path,
                    mimetype='text/csv',
                    as_attachment=True,
                    download_name='stock_import_template.csv'
                )
        else:  # xlsx
            template_path = os.path.join(template_dir, 'stock_template.xlsx')
            if os.path.exists(template_path):
                return send_file(
                    template_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='stock_import_template.xlsx'
                )

        # Fallback: Generate sample data if template files don't exist
        data = {
            'product_name': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Pen', 'Notebook'],
            'quantity': [10, 50, 30, 20, 100, 80],
            'rate': [50000.00, 500.00, 1500.00, 15000.00, 10.00, 25.00],
            'category': ['Electronics', 'Electronics', 'Electronics', 'Electronics', 'Stationery', 'Stationery'],
            'unit': ['pcs', 'pcs', 'pcs', 'pcs', 'pcs', 'pcs'],
            'low_stock_alert': [5, 10, 8, 5, 20, 15],
            'purchase_price': [40000.00, 350.00, 1200.00, 12000.00, 6.00, 18.00],
            'mrp': [55000.00, 599.00, 1799.00, 17999.00, 15.00, 35.00],
            'added_by_label': ['Ramesh', 'Priya', 'Ramesh', 'Priya', 'Ramesh', 'Priya'],
        }

        df = pd.DataFrame(data)
        output = io.BytesIO()

        if file_format == 'csv':
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name='stock_import_template.csv'
            )
        else:  # xlsx
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Stock Import')
            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='stock_import_template.xlsx'
            )

    except Exception as e:
        return jsonify({'error': 'Failed to generate template', 'message': str(e)}), 500


@stock_bp.route('/bulk-export', methods=['POST'])
@authenticate
@require_permission('export_stock')
def bulk_export_stock():
    """Export all stock data to CSV or Excel"""
    try:
        client_id = g.user['client_id']
        data = request.get_json() or {}
        file_format = data.get('format', 'csv')

        # Get all stock entries
        stock_entries = StockEntry.query.filter_by(client_id=client_id).order_by(StockEntry.product_name).all()

        if not stock_entries:
            return jsonify({'error': 'No stock data to export'}), 404

        # Convert to DataFrame
        data = []
        for stock in stock_entries:
            data.append({
                'product_name': stock.product_name,
                'quantity': stock.quantity,
                'rate': stock.rate,
                'category': stock.category,
                'unit': stock.unit,
                'low_stock_alert': stock.low_stock_alert,
                'created_at': stock.created_at.strftime('%Y-%m-%d %H:%M:%S') if stock.created_at else ''
            })

        df = pd.DataFrame(data)

        # Create in-memory file
        output = io.BytesIO()

        if file_format == 'csv':
            df.to_csv(output, index=False)
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name='stock_export.csv'
            )
        else:  # xlsx
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Stock')
            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='stock_export.xlsx'
            )

    except Exception as e:
        return jsonify({'error': 'Failed to export stock', 'message': str(e)}), 500


@stock_bp.route('/lookup/<code>', methods=['GET'])
@authenticate
def lookup_product(code):
    """
    Quick product lookup by barcode or item code
    Returns product with rate, GST%, HSN, current stock
    Perfect for barcode scanner integration
    """
    try:
        client_id = g.user['client_id']

        # Normalize input: strip whitespace and remove any extra characters
        normalized_code = code.strip() if code else ''

        if not normalized_code:
            return jsonify({
                'error': 'Product not found',
                'message': 'Empty search code provided'
            }), 404

        # Search by barcode first (most common for scanner)
        product = StockEntry.query.filter_by(
            client_id=client_id,
            barcode=normalized_code
        ).first()

        # Try barcode without spaces (some scanners add spaces)
        if not product:
            code_no_spaces = normalized_code.replace(' ', '')
            if code_no_spaces != normalized_code:
                product = StockEntry.query.filter_by(
                    client_id=client_id,
                    barcode=code_no_spaces
                ).first()

        # If not found by barcode, try item_code
        if not product:
            product = StockEntry.query.filter_by(
                client_id=client_id,
                item_code=normalized_code
            ).first()

        # If still not found, try exact product name match
        if not product:
            product = StockEntry.query.filter(
                StockEntry.client_id == client_id,
                StockEntry.product_name.ilike(normalized_code)
            ).first()

        if not product:
            return jsonify({
                'error': 'Product not found',
                'message': f'No product found with barcode, item code, or name: {code}'
            }), 404

        # Check stock availability (null-safe check for low_stock_alert)
        stock_status = 'available'
        if product.quantity == 0:
            stock_status = 'out_of_stock'
        elif product.low_stock_alert and product.quantity <= product.low_stock_alert:
            stock_status = 'low_stock'

        return jsonify({
            'success': True,
            'product': product.to_dict(),
            'stock_status': stock_status,
            'available_quantity': product.quantity
        }), 200

    except Exception as e:
        return jsonify({'error': 'Failed to lookup product', 'message': str(e)}), 500


@stock_bp.route('/export-low-stock', methods=['POST'])
@authenticate
@require_permission('export_stock')
def export_low_stock():
    """Export low stock items to PDF or Excel for easy ordering"""
    try:
        client_id = g.user['client_id']
        data = request.get_json() or {}
        file_format = data.get('format', 'xlsx')

        # Get low stock items
        low_stock = StockEntry.query.filter(
            StockEntry.client_id == client_id,
            StockEntry.quantity <= StockEntry.low_stock_alert
        ).order_by(StockEntry.quantity).all()

        if not low_stock:
            return jsonify({'error': 'No low stock items to export'}), 404

        # Prepare data for export
        export_data = []
        total_cost = 0
        for item in low_stock:
            need_to_order = max(0, item.low_stock_alert - item.quantity)
            estimated_cost = need_to_order * float(item.rate)
            total_cost += estimated_cost

            export_data.append({
                'Product Name': item.product_name,
                'Category': item.category or '-',
                'Current Stock': f"{item.quantity} {item.unit}",
                'Min. Required': f"{item.low_stock_alert} {item.unit}",
                'Need to Order': f"{need_to_order} {item.unit}",
                'Rate per Unit': f"₹{float(item.rate):.2f}",
                'Estimated Cost': f"₹{estimated_cost:.2f}"
            })

        df = pd.DataFrame(export_data)

        # Add total row
        total_row = {
            'Product Name': 'TOTAL',
            'Category': '',
            'Current Stock': '',
            'Min. Required': '',
            'Need to Order': '',
            'Rate per Unit': '',
            'Estimated Cost': f"₹{total_cost:.2f}"
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        output = io.BytesIO()

        if file_format == 'pdf':
            # For PDF, we'll create a simple HTML and convert it
            # You can use libraries like reportlab or weasyprint for better PDF generation
            html_content = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #dc2626; }}
                    table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #fee; color: #991b1b; font-weight: bold; }}
                    .total-row {{ background-color: #f0f0f0; font-weight: bold; }}
                    .header {{ margin-bottom: 20px; }}
                    .alert {{ background-color: #fee; padding: 15px; border-left: 4px solid #dc2626; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>[WARNING] Low Stock Report - Order Required</h1>
                    <p><strong>Generated on:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p><strong>Total Items:</strong> {len(low_stock)}</p>
                </div>
                <div class="alert">
                    <p><strong>Action Required:</strong> These items need immediate restocking to maintain inventory levels.</p>
                </div>
                {df.to_html(index=False, classes='table', escape=False)}
                <p style="margin-top: 20px;"><em>This is an automatically generated report from your billing system.</em></p>
            </body>
            </html>
            """

            # Simple HTML to PDF conversion (basic approach)
            # For production, use proper PDF libraries
            output.write(html_content.encode('utf-8'))
            output.seek(0)

            return send_file(
                output,
                mimetype='text/html',  # Change to 'application/pdf' with proper PDF library
                as_attachment=True,
                download_name=f'low_stock_report_{datetime.now().strftime("%Y%m%d")}.html'
            )
        else:  # Excel
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Low Stock Report')

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Low Stock Report']

                # Style the header
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                header_font = Font(bold=True, color='991B1B')

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left')

                # Style the total row
                last_row = worksheet.max_row
                total_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
                total_font = Font(bold=True)

                for cell in worksheet[last_row]:
                    cell.fill = total_fill
                    cell.font = total_font

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'low_stock_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )

    except Exception as e:
        return jsonify({'error': 'Failed to export low stock report', 'message': str(e)}), 500
